# ============================================================================
# Prométhée — Tests system_tools / sql_tools / python_tools / export_tools
# ============================================================================
import os, sys, json, tempfile, unittest
from pathlib import Path

_ROOT = Path(__file__).resolve().parent
for _ in range(5):
    if (_ROOT / "core" / "tools_engine.py").exists(): break
    _ROOT = _ROOT.parent
sys.path.insert(0, str(_ROOT))
os.environ.setdefault("MAX_CONTEXT_TOKENS", "128000")

from core.tools_engine import _TOOLS
_TOOLS.clear()
import tools.system_tools as _st_mod
import tools.sql_tools, tools.python_tools, tools.export_tools

from tools.system_tools import (
    read_file, write_file, head_file, tail_file, list_files,
    search_files, tree_view, copy_file, move_file, delete_file,
    create_directory, get_file_info, count_lines,
    find_and_replace, compress_files, extract_archive,
    diff_files, batch_rename, batch_delete,
)
from tools.sql_tools import (
    sql_connect, sql_disconnect, sql_list_connections,
    sql_list_tables, sql_describe, sql_query, sql_execute,
    sql_explain, sql_export_csv, _CONNECTIONS,
)
from tools.python_tools import python_exec, python_list_packages
from tools.export_tools import (
    export_md, export_xlsx_json, export_xlsx_csv,
    export_docx, export_pdf, export_pptx_outline,
)

# ── Patch système de fichiers ────────────────────────────────────────────────
# system_tools restreint les écritures à HOME (/root), inaccessible ici.
# On neutralise _is_safe_path pour les tests qui testent la LOGIQUE, pas la sécurité.
_SAFE_ORIG = _st_mod._is_safe_path
def _always_safe(path, operation="read"): return True, ""

class _SysBase(unittest.TestCase):
    """Patch _is_safe_path pour permettre écriture en /tmp et restaure après test."""
    def setUp(self):
        _st_mod._is_safe_path = _always_safe
        self.tmp = tempfile.mkdtemp()
    def tearDown(self):
        _st_mod._is_safe_path = _SAFE_ORIG
        import shutil; shutil.rmtree(self.tmp, ignore_errors=True)

# ══════════════════════════════════════════════════════════════════════════════
# SYSTEM TOOLS — lecture (pas de patch nécessaire)
# ══════════════════════════════════════════════════════════════════════════════

class TestReadFile(_SysBase):
    # Retourne: status, content, size, truncated, file
    def test_read_write(self):
        p = os.path.join(self.tmp,"t.txt")
        w = write_file(p,"Hello\nWorld\n"); self.assertEqual(w["status"],"success")
        r = read_file(p); self.assertIn("Hello",r["content"])
    def test_max_chars(self):
        p = os.path.join(self.tmp,"big.txt")
        write_file(p,"A"*10000)
        r = read_file(p,max_chars=100); self.assertLessEqual(len(r["content"]),200)
    def test_plage_lignes(self):
        p = os.path.join(self.tmp,"lines.txt")
        write_file(p,"\n".join(f"L{i}" for i in range(1,11)))
        r = read_file(p,start_line=3,end_line=5)
        self.assertIn("L3",r["content"]); self.assertNotIn("L1",r["content"])
    def test_ecriture_et_relecture(self):
        # write_file écrase toujours (mode='a' non implémenté)
        p = os.path.join(self.tmp,"a.txt")
        write_file(p,"contenu complet\n")
        r = read_file(p); self.assertIn("contenu complet",r["content"])
    def test_introuvable(self): self.assertEqual(read_file("/tmp/xyz_inexistant_abc.txt")["status"],"error")

class TestHeadTailFile(_SysBase):
    # head_file retourne: status, content, lines_shown, total_lines
    def setUp(self):
        super().setUp()
        self.p = os.path.join(self.tmp,"data.txt")
        write_file(self.p,"\n".join(f"L{i}" for i in range(1,21)))
    def test_head(self):
        r = head_file(self.p,lines=5)
        self.assertIn("L1",r["content"]); self.assertNotIn("L10",r["content"])
    def test_tail(self):
        r = tail_file(self.p,lines=3)
        self.assertIn("L20",r["content"]); self.assertNotIn("L2\n",r["content"])

class TestListSearchFiles(_SysBase):
    # list_files retourne: status, path, count, files (list of {name, path, type, size_kb, modified})
    def setUp(self):
        super().setUp()
        write_file(os.path.join(self.tmp,"a.txt"),"x")
        write_file(os.path.join(self.tmp,"b.py"),"print()")
        sub = os.path.join(self.tmp,"sub"); os.makedirs(sub,exist_ok=True)
        write_file(os.path.join(sub,"c.txt"),"cherche_moi")
    def test_list(self):
        r = list_files(self.tmp); noms = [f["name"] for f in r["files"]]
        self.assertIn("a.txt",noms); self.assertIn("b.py",noms)
    def test_pattern(self):
        r = list_files(self.tmp,pattern="*.txt")
        noms = [f["name"] for f in r["files"]]
        self.assertIn("a.txt",noms); self.assertNotIn("b.py",noms)
    def test_recursif(self):
        r = list_files(self.tmp,recursive=True)
        noms = [f["name"] for f in r["files"]]; self.assertIn("c.txt",noms)
    def test_search_nom(self):
        r = search_files(self.tmp,name_pattern="*.txt",recursive=True)
        self.assertGreater(len(r["results"]),0)
    def test_search_contenu(self):
        r = search_files(self.tmp,content="cherche_moi",recursive=True)
        self.assertGreater(len(r["results"]),0)

class TestTreeView(_SysBase):
    def test_basique(self):
        write_file(os.path.join(self.tmp,"f.txt"),"x")
        r = tree_view(self.tmp); self.assertIn("tree",r)

class TestCopyMoveDelete(_SysBase):
    # copy_file retourne: status, copied, to
    # move_file retourne: status, moved, to
    # delete_file retourne: status, deleted
    def setUp(self):
        super().setUp()
        self.src = os.path.join(self.tmp,"src.txt")
        write_file(self.src,"contenu source")
    def test_copy(self):
        dest = os.path.join(self.tmp,"copie.txt")
        r = copy_file(self.src,dest); self.assertEqual(r["status"],"success")
        self.assertTrue(Path(dest).exists()); self.assertTrue(Path(self.src).exists())
    def test_copy_sans_ecrasement(self):
        dest = os.path.join(self.tmp,"exist.txt"); write_file(dest,"x")
        r = copy_file(self.src,dest,overwrite=False); self.assertEqual(r["status"],"error")
    def test_copy_avec_ecrasement(self):
        dest = os.path.join(self.tmp,"exist2.txt"); write_file(dest,"x")
        r = copy_file(self.src,dest,overwrite=True); self.assertEqual(r["status"],"success")
    def test_move(self):
        dest = os.path.join(self.tmp,"moved.txt")
        r = move_file(self.src,dest); self.assertEqual(r["status"],"success")
        self.assertTrue(Path(dest).exists()); self.assertFalse(Path(self.src).exists())
    def test_delete_confirm(self):
        p = os.path.join(self.tmp,"del.txt"); write_file(p,"x")
        delete_file(p,confirm=True); self.assertFalse(Path(p).exists())
    def test_delete_sans_confirm(self):
        p = os.path.join(self.tmp,"nodel.txt"); write_file(p,"x")
        r = delete_file(p,confirm=False)
        self.assertNotEqual(r["status"],"success")  # cancelled ou error
        self.assertTrue(Path(p).exists())

class TestGetFileInfoCountLines(_SysBase):
    # get_file_info retourne: status, path, name, type, size, size_bytes, created, modified, permissions, md5
    def setUp(self):
        super().setUp()
        self.p = os.path.join(self.tmp,"t.txt")
        write_file(self.p,"a\nb\nc\n")
    def test_info(self):
        r = get_file_info(self.p); self.assertEqual(r["status"],"success")
        for k in ["size","modified","md5"]: self.assertIn(k,r)
    def test_count_lines(self):
        r = count_lines(self.tmp,pattern="*.txt",recursive=True)
        self.assertEqual(r["status"],"success"); self.assertIn("total_lines",r)

class TestFindAndReplace(_SysBase):
    def setUp(self):
        super().setUp()
        self.p = os.path.join(self.tmp,"doc.txt")
        write_file(self.p,"ancien_terme\nautre ancien_terme")
    def test_preview(self):
        r = find_and_replace(self.tmp,"ancien_terme","nouveau",pattern="*.txt",preview=True)
        self.assertEqual(r["mode"],"preview")
        self.assertIn("ancien_terme",Path(self.p).read_text())  # non modifié
    def test_apply(self):
        find_and_replace(self.tmp,"ancien_terme","nouveau",pattern="*.txt",preview=False)
        content = Path(self.p).read_text()
        self.assertIn("nouveau",content); self.assertNotIn("ancien_terme",content)

class TestCompressExtract(_SysBase):
    def setUp(self):
        super().setUp()
        self.f1 = os.path.join(self.tmp,"a.txt"); write_file(self.f1,"a")
        self.f2 = os.path.join(self.tmp,"b.txt"); write_file(self.f2,"b")
        self.arch = os.path.join(self.tmp,"arch.zip")
    def test_compress(self):
        r = compress_files([self.f1,self.f2],self.arch,format="zip")
        self.assertEqual(r["status"],"success"); self.assertTrue(Path(self.arch).exists())
    def test_extract(self):
        compress_files([self.f1],self.arch)
        dest = os.path.join(self.tmp,"ext")
        r = extract_archive(self.arch,destination=dest)
        self.assertEqual(r["status"],"success")
    def test_liste_seulement(self):
        compress_files([self.f1],self.arch)
        r = extract_archive(self.arch,liste_seulement=True)
        self.assertIn("entrees",r)  # clé réelle retournée par l'outil

class TestDiffFiles(_SysBase):
    # diff_files retourne: status, label_a, label_b, stats, diff, tronque
    def test_diff_texte_brut(self):
        r = diff_files("a\nb","a\nc",texte_brut=True)
        self.assertGreater(r["stats"]["modifiees"],0)
    def test_identiques(self):
        p = os.path.join(self.tmp,"s.txt"); write_file(p,"x")
        r = diff_files(p,p); self.assertTrue(r["stats"]["identique"])
    def test_fichiers(self):
        p1 = os.path.join(self.tmp,"v1.txt"); p2 = os.path.join(self.tmp,"v2.txt")
        write_file(p1,"L1\nL2\n"); write_file(p2,"L1\nMod\n")
        r = diff_files(p1,p2); self.assertFalse(r["stats"]["identique"])

class TestBatchOps(_SysBase):
    # batch_rename retourne: status, mode, renamed, results
    # batch_delete retourne: status, deleted, errors, deleted_files, error_files
    def setUp(self):
        super().setUp()
        for i in range(3): write_file(os.path.join(self.tmp,f"rapport_{i}.txt"),str(i))
    def test_batch_rename_preview(self):
        r = batch_rename(self.tmp,"rapport_","note_",pattern="*.txt",preview=True)
        self.assertEqual(r["mode"],"preview")
        self.assertTrue(any(Path(self.tmp).glob("rapport_*.txt")))
    def test_batch_rename_apply(self):
        r = batch_rename(self.tmp,"rapport_","note_",pattern="*.txt",preview=False)
        self.assertEqual(r["renamed"],3); self.assertTrue(any(Path(self.tmp).glob("note_*.txt")))
    def test_batch_delete_sans_confirm(self):
        files = [str(p) for p in Path(self.tmp).glob("*.txt")]
        r = batch_delete(files,confirm=False)
        self.assertNotEqual(r["status"],"success")  # cancelled ou error
    def test_batch_delete_avec_confirm(self):
        files = [str(p) for p in Path(self.tmp).glob("*.txt")]
        r = batch_delete(files,confirm=True)
        self.assertEqual(r["status"],"success"); self.assertFalse(any(Path(self.tmp).glob("*.txt")))

# ══════════════════════════════════════════════════════════════════════════════
# SQL TOOLS
# ══════════════════════════════════════════════════════════════════════════════

class _SqlBase(unittest.TestCase):
    URL = "sqlite:///:memory:"
    def setUp(self):
        _CONNECTIONS.clear()
        r = sql_connect(self.URL, nom="t")
        if r["status"] != "success": self.skipTest(f"SQLite indisponible: {r.get('error')}")
        sql_execute("CREATE TABLE e (id INTEGER PRIMARY KEY, nom TEXT, service TEXT, sal REAL)", connexion="t", confirmer=True)
        sql_execute("INSERT INTO e VALUES (1,'Dupont','RH',42000),(2,'Martin','IT',58000),(3,'Durand','RH',39000)", connexion="t", confirmer=True)
    def tearDown(self):
        try: sql_disconnect(nom="t")
        except: pass
        _CONNECTIONS.clear()

class TestSqlConnect(unittest.TestCase):
    def tearDown(self): _CONNECTIONS.clear()
    def test_sqlite(self):
        r = sql_connect("sqlite:///:memory:",nom="m"); self.assertEqual(r["status"],"success")
    def test_url_invalide(self):
        r = sql_connect("protocole_inconnu://localhost/db",nom="bad"); self.assertEqual(r["status"],"error")

class TestSqlQuery(_SqlBase):
    # sql_query retourne: status, connexion, sql, nb_lignes, tronque, colonnes, lignes, duree_ms
    def test_all(self):
        r = sql_query("SELECT * FROM e",connexion="t")
        self.assertEqual(r["status"],"success"); self.assertEqual(len(r["lignes"]),3)
    def test_where(self):
        r = sql_query("SELECT * FROM e WHERE service='RH'",connexion="t")
        self.assertEqual(len(r["lignes"]),2)
    def test_params(self):
        r = sql_query("SELECT * FROM e WHERE service=?",connexion="t",params=["IT"])
        self.assertEqual(len(r["lignes"]),1)
    def test_limite(self):
        r = sql_query("SELECT * FROM e",connexion="t",limite=1)
        self.assertLessEqual(len(r["lignes"]),1)
    def test_connexion_inexistante(self):
        self.assertEqual(sql_query("SELECT 1",connexion="fantome")["status"],"error")
    def test_sql_invalide(self):
        self.assertEqual(sql_query("SELECT FROMM erreur",connexion="t")["status"],"error")

class TestSqlExecute(_SqlBase):
    def test_insert(self):
        sql_execute("INSERT INTO e VALUES (4,'Petit','FIN',51000)",connexion="t",confirmer=True)
        r = sql_query("SELECT COUNT(*) as n FROM e",connexion="t")
        self.assertEqual(r["lignes"][0]["n"],4)
    def test_update(self):
        r = sql_execute("UPDATE e SET sal=45000 WHERE id=1",connexion="t",confirmer=True)
        self.assertEqual(r["status"],"success")
    def test_delete(self):
        sql_execute("DELETE FROM e WHERE id=3",connexion="t",confirmer=True)
        r = sql_query("SELECT COUNT(*) as n FROM e",connexion="t")
        self.assertEqual(r["lignes"][0]["n"],2)
    def test_sans_confirm(self):
        r = sql_execute("DELETE FROM e",connexion="t",confirmer=False)
        self.assertNotEqual(r["status"],"success")  # bloqué sans confirmer

class TestSqlMeta(_SqlBase):
    # sql_list_tables retourne: status, connexion, driver, nombre, tables [{nom, type, nb_lignes_approx}]
    # sql_describe retourne: status, connexion, table, nb_colonnes, colonnes, cles_primaires, cles_etrangeres
    # sql_list_connections retourne: status, nombre, connexions
    def test_list_tables(self):
        r = sql_list_tables(connexion="t"); self.assertEqual(r["status"],"success")
        self.assertIn("e",[t["nom"] for t in r["tables"]])
    def test_describe(self):
        r = sql_describe("e",connexion="t"); cols = [c["nom"] for c in r["colonnes"]]
        for c in ["id","nom","service","sal"]: self.assertIn(c,cols)
    def test_list_connections(self):
        r = sql_list_connections(); self.assertGreater(len(r["connexions"]),0)
    def test_disconnect(self):
        sql_disconnect(nom="t"); self.assertNotIn("t",_CONNECTIONS)

class TestSqlExplainExport(_SqlBase):
    # sql_explain retourne: status, connexion, driver, sql, plan, plan_texte
    # sql_export_csv retourne: status, fichier, ...
    def test_explain(self):
        r = sql_explain("SELECT * FROM e WHERE service='RH'",connexion="t")
        self.assertEqual(r["status"],"success"); self.assertIn("plan",r)
    def test_export_csv(self):
        tmp = tempfile.mkdtemp(); dest = os.path.join(tmp,"out.csv")
        r = sql_export_csv("SELECT * FROM e",connexion="t",destination=dest)
        self.assertEqual(r["status"],"success"); self.assertTrue(Path(dest).exists())
        lines = Path(dest).read_text().splitlines(); self.assertEqual(len(lines),4)  # header+3

# ══════════════════════════════════════════════════════════════════════════════
# PYTHON TOOLS
# ══════════════════════════════════════════════════════════════════════════════

class TestPythonExec(unittest.TestCase):
    # Retourne: status, output
    def test_simple(self):
        r = python_exec("print(2+2)"); self.assertEqual(r["status"],"success"); self.assertIn("4",r["output"])
    def test_import_stdlib(self):
        r = python_exec("import math; print(math.pi)"); self.assertIn("3.14",r["output"])
    def test_erreur_syntaxe(self):
        self.assertEqual(python_exec("def f(:")["status"],"error")
    def test_runtime_error(self):
        self.assertEqual(python_exec("x=1/0")["status"],"error")
    def test_timeout(self):
        self.assertEqual(python_exec("import time; time.sleep(60)",timeout=1)["status"],"error")
    def test_multiline(self):
        r = python_exec("a=10\nb=20\nprint(a+b)"); self.assertIn("30",r["output"])

class TestPythonListPackages(unittest.TestCase):
    # Retourne: status, count, packages
    def test_liste(self):
        r = python_list_packages(); self.assertEqual(r["status"],"success"); self.assertIn("count",r)

# ══════════════════════════════════════════════════════════════════════════════
# EXPORT TOOLS
# ══════════════════════════════════════════════════════════════════════════════
# Toutes ces fonctions retournent une str JSON avec status="ok" et path, size_bytes...

def _parse(r):
    """Normalise le retour (str JSON ou dict)."""
    if isinstance(r,str):
        try: return json.loads(r)
        except: return {"status":"error","error":r}
    return r

class TestExportMd(unittest.TestCase):
    def setUp(self): self.tmp = tempfile.mkdtemp()
    def test_creation(self):
        r = _parse(export_md("# Titre\nContenu.", output_path=os.path.join(self.tmp,"t.md")))
        self.assertEqual(r["status"],"ok"); self.assertTrue(Path(r["path"]).exists())
        self.assertIn("Titre",Path(r["path"]).read_text())
    def test_champs_retour(self):
        r = _parse(export_md("test",output_path=os.path.join(self.tmp,"t2.md")))
        for k in ["path","size_bytes","lines"]: self.assertIn(k,r)

class TestExportXlsx(unittest.TestCase):
    def setUp(self): self.tmp = tempfile.mkdtemp()
    def _wb(self): return {"sheets":[{"name":"S","headers":["A","B"],"rows":[[1,2],[3,4]]}]}
    def test_xlsx_json(self):
        r = _parse(export_xlsx_json(self._wb(),output_path=os.path.join(self.tmp,"t.xlsx")))
        self.assertEqual(r["status"],"ok"); self.assertTrue(Path(r["path"]).exists())
        import openpyxl; ws = openpyxl.load_workbook(r["path"])["S"]
        self.assertEqual(ws.cell(1,1).value,"A"); self.assertEqual(ws.cell(2,1).value,1)
    def test_xlsx_csv(self):
        r = _parse(export_xlsx_csv("a,b\n1,2\n3,4\n",output_path=os.path.join(self.tmp,"t2.xlsx")))
        self.assertEqual(r["status"],"ok"); self.assertTrue(Path(r["path"]).exists())
    def test_champs_retour(self):
        r = _parse(export_xlsx_json(self._wb(),output_path=os.path.join(self.tmp,"t3.xlsx")))
        for k in ["path","size_bytes","total_rows"]: self.assertIn(k,r)

class TestExportDocx(unittest.TestCase):
    def setUp(self): self.tmp = tempfile.mkdtemp()
    def test_creation(self):
        doc = {"title":"T","sections":[{"heading":"H","content":"C contenu."}]}
        r = _parse(export_docx(doc,output_path=os.path.join(self.tmp,"t.docx")))
        self.assertEqual(r["status"],"ok"); self.assertTrue(Path(r["path"]).exists())
        self.assertGreater(r["size_bytes"],0)

class TestExportPdf(unittest.TestCase):
    def setUp(self): self.tmp = tempfile.mkdtemp()
    def test_creation(self):
        doc = {"title":"PDF","sections":[{"heading":"H","content":"Contenu PDF."}]}
        r = _parse(export_pdf(doc,output_path=os.path.join(self.tmp,"t.pdf")))
        self.assertEqual(r["status"],"ok")
        data = Path(r["path"]).read_bytes(); self.assertTrue(data.startswith(b"%PDF"))

class TestExportPptx(unittest.TestCase):
    def setUp(self): self.tmp = tempfile.mkdtemp()
    def test_outline(self):
        r = _parse(export_pptx_outline("Titre\n- Slide 1\n- Slide 2",title="Prez",
                                       output_path=os.path.join(self.tmp,"t.pptx")))
        self.assertEqual(r["status"],"ok"); self.assertTrue(Path(r["path"]).exists())
        self.assertGreater(r["slides"],0)

if __name__ == "__main__":
    unittest.main(verbosity=2)
